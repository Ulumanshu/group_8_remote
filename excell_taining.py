import openpyxl
from xlsxwriter.workbook import Workbook
from datetime import datetime
from pprint import pprint


def get_xlsx_sheet_data(xlsx_file_path):
    wb_obj = openpyxl.load_workbook(xlsx_file_path)
    sheet = wb_obj.active
    res = []
    for row_ in sheet.iter_rows():
        row_values = []
        for cell in row_:
            row_values.append(cell.value)

        res.append(row_values)

    return res


if __name__ == "__main__":
    xlsx_file = '/home/wooden/workspace/sdacademy/group_8_remote/my_file.xlsx'
    excell_data = get_xlsx_sheet_data(xlsx_file)
    pprint(excell_data, indent=4)

    # workbook = Workbook('result.xlsx')
    with Workbook('result.xlsx') as workbook:
        worksheet = workbook.add_worksheet(str('Sheet1'))
        for row, data_line in enumerate(excell_data):
            for col, data_value in enumerate(data_line):
                if isinstance(data_value, datetime):
                    data_value = data_value.strftime("%Y - %m - %d")
                if isinstance(data_value, float):
                    data_value = round(data_value, 2)

                worksheet.write(row, col, data_value)  # Writes

        chart_pie = workbook.add_chart({'type': 'column'})
        # chart_pie.add_series({'values': '=Sheet1!$B$1:$B$5'})
        # chart_pie.add_series({'values': '=Sheet1!$C$1:$C$5'})
        chart_pie.add_series({
            'categories': ['Sheet1', 1, 0, 5, 0],
            'values': ['Sheet1', 1, 4, 5, 4],
            'data_labels': {
                'legend_key': True,
                'category': True,
                'value': True,
                'percentage': True
            },
            'name': 'Belekas'
        })
        chart_pie.set_title({'name': 'Graph1'})
        chart_pie.set_style(10)
        worksheet.insert_chart("B10", chart_pie, {'x_scale': 2, 'y_scale': 2})

    # workbook.close()

    # varied_array = [1, 2.99, 'text', (1, 9), [1, 2], {'names': 200000}, {1, 2, 3}, None, True, False, datetime.strptime('2011-04-01', '%Y-%m-%d')]
    # for nr, e in enumerate(varied_array):
    #
    #     # print('ELEMENTAS', e)
    #     # print("ELEMENTO TIPAS", type(e))
    #     # print("ISINSTACE FLOAT/INT/DATE?: ", isinstance(e, (float, datetime)))
    #     if isinstance(e, (float, int)):
    #         print(e)
    #         print(nr)
    #         print(nr + e)
    #         print('___________')
